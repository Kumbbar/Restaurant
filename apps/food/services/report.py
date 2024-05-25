import datetime

from django.http import HttpResponse
from openpyxl.workbook import Workbook

from apps.food.models import OrderDish, Dish, Order


def get_restaurant_excel_report(request):
    data = {
        'restaurant_id': request.GET.get('restaurant_id'),
        'time_of_start': request.GET.get('time_of_start'),
        'time_of_end': request.GET.get('time_of_end')
    }
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="restaurant-report.xlsx"'

    wb = Workbook()
    del wb['Sheet']

    ws = wb.create_sheet('popular dishes')
    headers = ["name", "price", "type", "count", "total_price"]
    ws.append(headers)
    order_dishes_data = OrderDish.objects.raw('''
                with all_data as (
                    select
                        1 as id,
                        food_dish.name,
                        food_dish.price,
                        fd.name as type,
                        sum(food_orderdish.count) as count,
                        sum(food_orderdish.count) * food_dish.price as total_price

                    from food_dish
                    join food_orderdish on food_dish.id = food_orderdish.dish_id
                    join food_order on food_orderdish.order_id = food_order.id
                    join food_dishtype fd on fd.id = food_dish.dish_type_id
                    where food_order.restaurant_id = %(restaurant_id)s and date(food_order.created_at) BETWEEN date(%(time_of_start)s) and date(%(time_of_end)s)
                    group by food_dish.name
                    order by 6 desc
                )

                select * from all_data
                union all
                select
                    1 as id,
                    '' as name,
                    '' as price,
                    '' as type,
                    sum(count) as count,
                    sum(total_price) as total_price
                    from all_data

                ''', data)

    for product in order_dishes_data:
        ws.append([product.name, product.price, product.type, product.count, product.total_price])

    ws = wb.create_sheet('orders')
    headers = ["name", "surname", "patronymic", "created_at", "total_price"]
    ws.append(headers)
    orders = OrderDish.objects.raw('''
                    with all_data as (
                        select
                            1 as id,
                            fc.name,
                            fc.surname,
                            fc.patronymic,
                            food_order.created_at,
                            sum(food_orderdish.count * food_dish.price) as total_price
                        from food_dish
                        join food_orderdish on food_dish.id = food_orderdish.dish_id
                        join food_order on food_orderdish.order_id = food_order.id
                        left join main.food_client fc on food_order.client_id = fc.id
                        where food_order.restaurant_id = %(restaurant_id)s and date(food_order.created_at) BETWEEN date(%(time_of_start)s) and date(%(time_of_end)s)
                        group by fc.name, fc.surname, fc.patronymic, food_order.created_at
                        order by 5, 6 desc
                    )
                    select * from all_data
                    union all
                    select
                        1 as id,
                        '' as name,
                        '' as surname,
                        '' as patronymic,
                        '' as created_at,
                        sum(total_price) as total_price
                        from all_data
                    ''', data)

    for order in orders:
        ws.append([order.name, order.surname, order.patronymic, order.created_at, order.total_price])

    ws = wb.create_sheet('clients')
    headers = ["name", "surname", "patronymic", "total_price"]
    ws.append(headers)
    orders = OrderDish.objects.raw('''
                        with all_data as (
                            select
                                1 as id,
                                fc.name,
                                fc.surname,
                                fc.patronymic,
                                sum(food_orderdish.count * food_dish.price) as total_price
                            from food_dish
                            join food_orderdish on food_dish.id = food_orderdish.dish_id
                            join food_order on food_orderdish.order_id = food_order.id
                            join main.food_client fc on food_order.client_id = fc.id
                            where food_order.restaurant_id = %(restaurant_id)s and date(food_order.created_at) BETWEEN date(%(time_of_start)s) and date(%(time_of_end)s)
                            group by fc.name, fc.surname, fc.patronymic
                            order by 5 desc
                        )
                        select * from all_data
                        union all
                        select
                            1 as id,
                            '' as name,
                            '' as surname,
                            '' as patronymic,
                            sum(total_price) as total_price
                            from all_data
                        ''', data)

    for order in orders:
        ws.append([order.name, order.surname, order.patronymic, order.total_price])


    wb.save(response)
    return response